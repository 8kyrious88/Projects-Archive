import openpyxl, pandas, datetime

date = datetime.datetime.now()
date = date.strftime("%d%m")

shopee_file = rf"C:\Users\Ng ChunPing\Desktop\PDF Testing\Shopee {date}25.xlsx"
lzd_file = rf"C:\Users\Ng ChunPing\Desktop\PDF Testing\Lzd {date}25.xlsx"
consolidated_file = r"C:\Users\Ng ChunPing\Desktop\PDF Testing\Calculated.csv"

shopee_wb = openpyxl.load_workbook(shopee_file)
shopee_orders = shopee_wb["orders"]
lzd_wb = openpyxl.load_workbook(lzd_file)
lzd_orders = lzd_wb["sheet1"]
all_product_name = []
quantity = []
items_required = {}

temp_dict = {}

for order in shopee_orders.iter_rows(min_row=1, min_col=12, max_col=12, values_only=True):
    for product_name in order:
        if product_name == "Product Name":
            pass
        else:
            all_product_name.append(product_name)

for order in lzd_orders.iter_rows(min_row=1, min_col=52, max_col=52, values_only=True):
    for product_name in order:
        if product_name == "itemName":
            pass
        else:
            all_product_name.append(product_name)

for order in shopee_orders.iter_rows(min_row=1, min_col=17, max_col=17, values_only=True):
    for qty in order:
        if qty == "Quantity":
            pass
        else:
            quantity.append(qty)

while len(quantity) < len(all_product_name):
    quantity.append(1)


for num in range(len(all_product_name)):
    if all_product_name[num] in items_required:
        items_required[all_product_name[num]] += int(quantity[num])
    else:
        items_required[all_product_name[num]] = int(quantity[num])

temp_list_1 = [key for (key,value) in items_required.items()]
temp_list_2 = [value for (key,value) in items_required.items()]

temp_dict["Product Name"] = temp_list_1
temp_dict["Quantity"] = temp_list_2

df = pandas.DataFrame(temp_dict)
df.to_csv(consolidated_file)

shopee_wb.close()
lzd_wb.close()
print("Finished")